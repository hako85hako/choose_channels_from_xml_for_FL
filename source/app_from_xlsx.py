from glob import glob
import csv
import traceback
import pandas as pd
import re
import openpyxl

import frame


def main(old_xlsx,new_xlsx,target_xlsm):
    # 初期設定 ####################################################
    # old_flg
    first_flg = 0

    # 列指定offset
    row_offset = 12

    # データのあるシートを指定
    pattern_1 = r'\'【SQL】tbl_datarelation\(構造\)\''
    repatter_1 = re.compile(pattern_1)
    
    # データのカラム名を指定
    column_1 = 'comment'
    column_2 = 'columnno'
    column_3 = 'slope'
    column_4 = 'intercept'

    # IDとpasswordの情報があるシートを指定 
    pattern_2 = r'\'【SQL】tbl_comdevicies\''
    repatter_2 = re.compile(pattern_2)
    
    # データのカラム名を指定
    column_5 = 'comid'
    column_6 = 'password'

    #新旧サイト判定
    before = 'before'
    after = 'after'

    # Exception_text
    not_found_target_file = '指定されたファイルがありません'
    not_correct_target_file = '正しい.xlsxファイルを指定してください'
    not_found_write_file = '書き込む.xlsmファイルがありません'
    write_file_two_or_more = '書き込み対象の.xlsmファイルが2個以上存在します'
    save_failed_error = '保存処理に失敗しました。\n指定したExcelを開いていないか確認してください。'
    unexpected_error = '予期せぬエラーです。管理者にご相談ください。'
    #############################################################
    try:
        #ファイル指定
        target_xlsx_names = [old_xlsx,new_xlsx]

        for target_xlsx_name in target_xlsx_names:
            #初回（旧サイトファイル判定用インクリメント）
            first_flg += 1
            #ファイルの存在判定
            if target_xlsx_name == "選択をキャンセルしました" or target_xlsx_name == "未選択です":
                continue
            #オフセットデータ初期化
            offset_data = 0
            #target_sheetの有無確認用
            count_target_sheet = 0
            try:
                #sheet_namesメソッドでExcelブック内の各シートの名前をリストで取得できる
                input_sheet_names = pd.ExcelFile(target_xlsx_name).sheet_names
            except:
                #ファイルが存在しない場合error
                frame.show_error(not_found_target_file)
                return False
            #格納先作成
            datas = dict()
            
            for input_sheet_name in input_sheet_names:
                #もしpattern_1にmatchするなら読み込み開始
                result = repatter_1.match(repr(input_sheet_name))
                if result:
                    count_target_sheet += 1
                    #dict型で格納
                    df_sheet_multi = pd.read_excel(target_xlsx_name, sheet_name=[input_sheet_name])
                    #index番号作成
                    index = 0 
                    #dataの行数を求める
                    lens = df_sheet_multi[input_sheet_name].shape[0]
                    for len_data in range(lens):
                        data = dict()
                        data[column_1] = df_sheet_multi[input_sheet_name][column_1][index]
                        data[column_2] = df_sheet_multi[input_sheet_name][column_2][index]
                        data[column_3] = df_sheet_multi[input_sheet_name][column_3][index]
                        data[column_4] = df_sheet_multi[input_sheet_name][column_4][index]
                        datas[data[column_2]] = data
                        index += 1
                    #入れ替え後と入れ替え前の指定
                    get_datas = dict()
                    if first_flg == 1:
                        #ch順にソート
                        get_datas[before] = sorted(datas.items())
                    else :
                        #ch順にソート
                        get_datas[after] = sorted(datas.items())
            if count_target_sheet == 0:
                frame.show_error(not_correct_target_file)
                return False

            # 書き込み用.xlsm指定
            target_xlsm_name = target_xlsm
            if not target_xlsm_name:
                frame.show_error(not_found_write_file)
                return False

            #書き込み用のExcel指定
            # ワークブック作成
            wb = openpyxl.Workbook()
            # ワークブックの読み込み
            wb = openpyxl.load_workbook(target_xlsm_name,keep_vba=True)
            # 読み込んだブックのシート選択
            sheet = wb.worksheets[0]

            # 書き込み処理
            index = 0 
            if before in get_datas:
                for get_data in get_datas[before]:
                    if index == 0:
                        offset_data = get_data[1][column_2]
                    index += 1
                    sheet.cell(column=2, row=(row_offset+index), value=str(index))
                    sheet.cell(column=3, row=(row_offset+index), value=get_data[1][column_1])
                    sheet.cell(column=12, row=(row_offset+index), value=round(get_data[1][column_3],3))
                    sheet.cell(column=13, row=(row_offset+index), value=round(get_data[1][column_4],3))
                sheet.cell(column=12, row=11, value=offset_data)
                try:
                    # 保存
                    wb.save(target_xlsm_name)
                except:
                    # 保存処理が失敗した場合
                    frame.show_error(save_failed_error)
                    return False

            elif after in get_datas:          
                #xls book Open (xls, xlsxのどちらでも可能)
                input_book = pd.ExcelFile(target_xlsx_name)
                
                
                #sheet_namesメソッドでExcelブック内の各シートの名前をリストで取得できる
                input_sheet_names = input_book.sheet_names
                FL_ID = ""
                password = ""
                for input_sheet_name in input_sheet_names:
                    result = repatter_2.match(repr(input_sheet_name))
                    #もしpatternにmatchするなら読み込み開始
                    if result:
                        #dict型で格納
                        df_sheet_multi = pd.read_excel(target_xlsx_name, sheet_name=[input_sheet_name])
                        FL_ID = df_sheet_multi[input_sheet_name][column_5][0]
                        password = df_sheet_multi[input_sheet_name][column_6][0]    

                for get_data in get_datas[after]:
                    if index == 0:
                        offset_data = get_data[1][column_2]
                    index += 1
                    sheet.cell(column=15, row=(row_offset+index), value=str(index))
                    sheet.cell(column=16, row=(row_offset+index), value=get_data[1][column_1])
                    sheet.cell(column=24, row=(row_offset+index), value=round(get_data[1][column_3],3))
                    sheet.cell(column=25, row=(row_offset+index), value=round(get_data[1][column_4],3))
                sheet.cell(column=5, row=2, value=FL_ID)
                sheet.cell(column=12, row=2, value=password)
                sheet.cell(column=24, row=11, value=offset_data)
                
                try:
                    # 保存
                    wb.save(target_xlsm_name)
                except:
                    # 保存処理が失敗した場合
                    frame.show_error(save_failed_error)
                    return False
            else:
                frame.show_error(unexpected_error)
                return False
        return True
    except:
        ex = traceback.format_exc()
        print(ex)
        frame.TkinterClass.call_except_in_app(ex)
        return False

#テスト用
# if __name__=="__main__":
#    main()